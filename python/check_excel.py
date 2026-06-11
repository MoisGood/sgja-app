import openpyxl
import os

dir_path = r"C:\Proyectos\BackUp\python"
files = ["User_Download_09062026_202209_File.xlsx", "User_Download_09062026_202122_File.xlsx"]

for fname in files:
    path = os.path.join(dir_path, fname)
    if not os.path.exists(path):
        print(f"{fname}: NO EXISTE")
        continue
    
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    print(f"\n=== {fname} ===")
    print(f"Filas: {ws.max_row}, Columnas: {ws.max_column}")
    
    # Mostrar header + primeras 3 filas
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        print(f"  Fila {i+1}: {list(row)[:15]}")
        if i >= 3:
            break
    wb.close()
